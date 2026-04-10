from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

from .phase1_thermal import MAX_CONCENTRATION_CM3


@dataclass(slots=True)
class MeasuredInitialProfile:
    depth_nm: np.ndarray
    total_p_cm3: np.ndarray
    active_p_cm3: np.ndarray
    inactive_p_cm3: np.ndarray


def _find_ecv_header(rows: list[list[str]]) -> int:
    for idx, row in enumerate(rows):
        if "Depth" in row and "N(1/cm" in ",".join(row):
            return idx
    raise ValueError("Could not locate the ECV data header row.")


def _parse_float(value: str) -> float | None:
    text = value.strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_ecv_raw_csv(path: str | Path) -> tuple[np.ndarray, np.ndarray]:
    with Path(path).open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        rows = list(csv.reader(handle))

    header_idx = _find_ecv_header(rows)
    header = rows[header_idx]
    depth_idx = header.index("Depth")
    active_idx = next(idx for idx, cell in enumerate(header) if cell.startswith("N(1/cm"))

    depth_um: list[float] = []
    active_cm3: list[float] = []
    for row in rows[header_idx + 1 :]:
        if depth_idx >= len(row) or active_idx >= len(row):
            continue
        depth_value = _parse_float(row[depth_idx])
        active_value = _parse_float(row[active_idx])
        if depth_value is None or active_value is None:
            continue
        depth_um.append(depth_value)
        active_cm3.append(active_value)

    if not depth_um:
        raise ValueError("No numeric ECV profile rows were found.")

    return np.asarray(depth_um, dtype=float) * 1.0e3, np.asarray(active_cm3, dtype=float)


def load_sims_raw_xlsx(path: str | Path, location: str = "CTV") -> tuple[np.ndarray, np.ndarray]:
    workbook = load_workbook(Path(path), data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    location = location.strip().upper()
    block_marker = "#C" if location.startswith("C") else "#J"

    block_col = None
    for cell in sheet[2]:
        if cell.value == block_marker:
            block_col = cell.column
            break
    if block_col is None:
        raise ValueError(f"Could not find the SIMS block marker {block_marker!r}.")

    depth_col = None
    total_col = None
    for offset in range(7):
        x_header = sheet.cell(row=4, column=block_col + offset).value
        y_header = sheet.cell(row=4, column=block_col + offset + 1).value
        if x_header == "31P(X)" and y_header == "31P(Y)":
            depth_col = block_col + offset
            total_col = block_col + offset + 1
            break
    if depth_col is None or total_col is None:
        raise ValueError(f"Could not find the 31P(X/Y) columns in the SIMS block {block_marker!r}.")

    depth_nm: list[float] = []
    total_cm3: list[float] = []
    for row in sheet.iter_rows(min_row=6, values_only=True):
        depth_value = row[depth_col - 1]
        total_value = row[total_col - 1]
        if depth_value is None or total_value is None:
            continue
        depth_nm.append(float(depth_value))
        total_cm3.append(float(total_value))

    if not depth_nm:
        raise ValueError("No numeric SIMS profile rows were found.")

    return np.asarray(depth_nm, dtype=float), np.asarray(total_cm3, dtype=float)


def _sort_and_unique(depth_nm: np.ndarray, values_cm3: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    order = np.argsort(depth_nm)
    depth_sorted = np.asarray(depth_nm[order], dtype=float)
    value_sorted = np.asarray(values_cm3[order], dtype=float)
    unique_depth, unique_index = np.unique(depth_sorted, return_index=True)
    return unique_depth, value_sorted[unique_index]


def _prepend_surface_point(depth_nm: np.ndarray, values_cm3: np.ndarray) -> tuple[np.ndarray, np.ndarray]:
    depth_sorted, value_sorted = _sort_and_unique(depth_nm, values_cm3)
    if depth_sorted[0] <= 0.0:
        return depth_sorted, value_sorted

    if depth_sorted.size >= 2:
        x0, x1 = depth_sorted[0], depth_sorted[1]
        y0 = np.log10(np.clip(value_sorted[0], 1.0, MAX_CONCENTRATION_CM3))
        y1 = np.log10(np.clip(value_sorted[1], 1.0, MAX_CONCENTRATION_CM3))
        slope = (y1 - y0) / max(x1 - x0, 1.0e-12)
        surface_log = y0 - slope * x0
        surface_value = float(np.clip(10.0**surface_log, 1.0, MAX_CONCENTRATION_CM3))
    else:
        surface_value = float(np.clip(value_sorted[0], 1.0, MAX_CONCENTRATION_CM3))

    return (
        np.concatenate(([0.0], depth_sorted)),
        np.concatenate(([surface_value], value_sorted)),
    )


def _nonincreasing_envelope(values_cm3: np.ndarray) -> np.ndarray:
    return np.minimum.accumulate(np.asarray(values_cm3, dtype=float))


def interpolate_profile_log_cm3(
    depth_target_nm: np.ndarray,
    depth_source_nm: np.ndarray,
    values_source_cm3: np.ndarray,
) -> np.ndarray:
    source_depth, source_values = _prepend_surface_point(depth_source_nm, values_source_cm3)
    log_values = np.log10(np.clip(source_values, 1.0, MAX_CONCENTRATION_CM3))
    interpolated_log = np.interp(
        np.asarray(depth_target_nm, dtype=float),
        source_depth,
        log_values,
        left=log_values[0],
        right=log_values[-1],
    )
    return np.clip(10.0**interpolated_log, 0.0, MAX_CONCENTRATION_CM3)


def build_measured_initial_profile(
    ecv_csv_path: str | Path,
    sims_xlsx_path: str | Path,
    sims_location: str = "CTV",
    depth_grid_nm: np.ndarray | None = None,
) -> MeasuredInitialProfile:
    ecv_depth_nm, ecv_active_cm3 = load_ecv_raw_csv(ecv_csv_path)
    sims_depth_nm, sims_total_cm3 = load_sims_raw_xlsx(sims_xlsx_path, location=sims_location)

    if depth_grid_nm is None:
        depth_grid_nm = np.unique(
            np.concatenate(
                [
                    _prepend_surface_point(ecv_depth_nm, ecv_active_cm3)[0],
                    _prepend_surface_point(sims_depth_nm, sims_total_cm3)[0],
                ]
            )
        )
    depth_grid_nm = np.asarray(depth_grid_nm, dtype=float)
    total_p_cm3 = interpolate_profile_log_cm3(depth_grid_nm, sims_depth_nm, sims_total_cm3)
    active_p_cm3 = interpolate_profile_log_cm3(depth_grid_nm, ecv_depth_nm, ecv_active_cm3)
    total_p_cm3 = _nonincreasing_envelope(total_p_cm3)
    active_p_cm3 = _nonincreasing_envelope(active_p_cm3)
    active_p_cm3 = np.minimum(active_p_cm3, total_p_cm3)
    inactive_p_cm3 = np.maximum(total_p_cm3 - active_p_cm3, 0.0)
    return MeasuredInitialProfile(
        depth_nm=depth_grid_nm,
        total_p_cm3=total_p_cm3,
        active_p_cm3=active_p_cm3,
        inactive_p_cm3=inactive_p_cm3,
    )


def load_measured_initial_profile_csv(path: str | Path) -> MeasuredInitialProfile:
    array = np.genfromtxt(Path(path), delimiter=",", names=True)
    return MeasuredInitialProfile(
        depth_nm=np.asarray(array["depth_nm"], dtype=float),
        total_p_cm3=np.asarray(array["total_p_cm3"], dtype=float),
        active_p_cm3=np.asarray(array["active_p_cm3"], dtype=float),
        inactive_p_cm3=np.asarray(array["inactive_p_cm3"], dtype=float),
    )


def save_measured_initial_profile_csv(profile: MeasuredInitialProfile, path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    data = np.column_stack(
        [
            profile.depth_nm,
            profile.total_p_cm3,
            profile.active_p_cm3,
            profile.inactive_p_cm3,
        ]
    )
    np.savetxt(
        output_path,
        data,
        delimiter=",",
        header="depth_nm,total_p_cm3,active_p_cm3,inactive_p_cm3",
        comments="",
    )
    return output_path


def save_measured_profile_plot(profile: MeasuredInitialProfile, path: str | Path) -> Path:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    figure, axis = plt.subplots(figsize=(8.2, 4.8))
    axis.semilogy(profile.depth_nm, np.maximum(profile.total_p_cm3, 1.0e10), color="#c0392b", lw=2.0, label="SIMS total P")
    axis.semilogy(profile.depth_nm, np.maximum(profile.active_p_cm3, 1.0e10), color="#2471a3", lw=2.0, label="ECV active P")
    axis.semilogy(
        profile.depth_nm,
        np.maximum(profile.inactive_p_cm3, 1.0e10),
        color="#7d6608",
        lw=1.8,
        ls="--",
        label="Inactive P = max(SIMS - ECV, 0)",
    )
    axis.set_xlabel("Depth (nm)")
    axis.set_ylabel("Concentration (cm^-3)")
    axis.set_title("Measured Initial Profiles Used by the Model")
    axis.grid(alpha=0.25)
    axis.legend()
    figure.tight_layout()
    figure.savefig(output_path, dpi=220)
    plt.close(figure)
    return output_path
